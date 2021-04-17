# Author: Neaz Morshed

# Date: 28/10/2020

from time import sleep
from tkinter import filedialog
from tkinter import *
import tkinter as tk
import openpyxl
from tkinter import StringVar
from tkinter import messagebox


# Calculate t-Distribution
def calTDis():
    path = open_browse_excel_window()
    # txtdissr.set("Hello")
    try:
        workbook = openpyxl.load_workbook(path, data_only=True)
        sheets = workbook.sheetnames
        print(sheets, len(sheets))
        tsheet = workbook[sheets[0]]
        print(tsheet)
        tsheet['I4'] = tdistvalue.get()
        tsheet['I5'] = tdisdfvalue.get()
        workbook.save(path)
        val1 = tsheet.cell(column=9, row=6).value
        val2 = tsheet.cell(column=9, row=7).value
        val3 = tsheet.cell(column=9, row=8).value
        txtdissr.set(val1)
        tnetprob.set(val2)
        tprobdiff.set(val3)
    except FileNotFoundError:
        messagebox._show("File not found","File 'Sample_DistributionCalculator.xlsx' not found in same Folder/Directory")


# Calculate Chi-Square Distribution
def calChiDis():
    path = open_browse_excel_window()
    # txtdissr.set("Hello")
    try:
        workbook = openpyxl.load_workbook(path, data_only=True)
        sheets = workbook.sheetnames
        print(sheets, len(sheets))
        chisheet = workbook[sheets[1]]
        print(chisheet)
        chisheet['I4'] = tdistvalue2.get()
        chisheet['I5'] = tdisdfvalue2.get()
        workbook.save(path)
        val1 = chisheet.cell(column=9, row=6).value
        val2 = chisheet.cell(column=9, row=7).value
        val3 = chisheet.cell(column=9, row=8).value
        x2.set(val1)
        y2.set(val2)
        z2.set(val3)
    except FileNotFoundError:
        messagebox._show("File not found","File 'Sample_DistributionCalculator.xlsx' not found in same Folder/Directory")

# Calculate Normal Distribution
def calNorDis():
    path = open_browse_excel_window()
    # txtdissr.set("Hello")
    try:
        workbook = openpyxl.load_workbook(path, data_only=True)
        sheets = workbook.sheetnames
        print(sheets, len(sheets))
        chisheet = workbook[sheets[2]]
        print(chisheet)
        chisheet['I4'] = tdistvalue3.get()
        workbook.save(path)
        val1 = chisheet.cell(column=9, row=5).value
        val2 = chisheet.cell(column=9, row=6).value
        val3 = chisheet.cell(column=9, row=7).value
        print(x3)
        x3.set(val1)
        y3.set(val2)
        z3.set(val3)
    except FileNotFoundError:
        messagebox._show("File not found","File 'Sample_DistributionCalculator.xlsx' not found in same Folder/Directory")

# Returns the Excel File
def open_browse_excel_window():
    return "Sample_DistributionCalculator.xlsx"


# Making selections
def selcal():
    selvar1 = str(selvar.get())
    if(selvar1=="1"):
        visframe1.place(x=280, y=50, width=690, height=380)
        visframe2.place_forget()
        visframe3.place_forget()
        selLabel.place_forget()

    elif(selvar1=="2"):
        visframe2.place(x=280, y=50, width=690, height=380)
        visframe1.place_forget()
        visframe3.place_forget()
        selLabel.place_forget()

    elif(selvar1=="3"):
        visframe3.place(x=280, y=50, width=690, height=380)
        visframe2.place_forget()
        visframe1.place_forget()
        selLabel.place_forget()

# Navigation from Title to Main screen
def Start():
    TitleFrame.place_forget()
    Frame2.place(width=1000,height=500)

# creating Main window
root = Tk()
root.title("Calculator")
root.minsize(1000, 500)
root.maxsize(1000, 500)

visiframe2bg = "#d84315"
visiframe3bg = "#d84315"

TitleFrame = Frame(root,bg="#c41c00")

Frame2 = Frame(root,bg="#bf360c")
sideframe = Frame(Frame2,bg="#870000")

visframe1 = Frame(Frame2,bg="#d84315")
visframe2 = Frame(Frame2,bg=visiframe2bg)
visframe3 = Frame(Frame2,bg=visiframe3bg)


# Frame TitleFrame for Creating the Title Page
label_heading = tk.Label(TitleFrame,text="Statistical Distribution and Interpreting P values Calculator",font=("Gabriola", 40),bg="#c41c00",fg="white").place(x=0,y=50,width=1000)
button_start = tk.Button(TitleFrame,text="Lets Calculate",font=("Times",15),bg="white",fg="black",command=Start).place(x=425,y=200)


# Frame sideframe

selvar = IntVar()

# Radio Buttons
chk1 = tk.Radiobutton(sideframe,selectcolor='#3e2723', text="t-Distribution",font=("Times",15,"bold"),fg="#fff3e0",bg="#870000",variable=selvar,value=1,command=selcal).grid(row=4,column=0,sticky=tk.W,padx=10,pady=95)
chk2 = tk.Radiobutton(sideframe,selectcolor='#3e2723', text="Chi-square Distribution",font=("Times",15),fg="#fff3e0",bg="#870000",variable=selvar,value=2,command=selcal).grid(row=5,column=0,sticky=tk.W,padx=10)
chk3 = tk.Radiobutton(sideframe,selectcolor='#3e2723', text="Normal Distribution",font=("Times",15),fg="#fff3e0",bg="#870000",variable=selvar,value=3,command=selcal).grid(row=6,column=0,columnspan=1,sticky=tk.W,padx=10,pady=95)

sideframe.place(x=0,y=0,width=250,height=500)

# Frame2 and its subframes
selLabel = Label(Frame2,text = "Select a Calculator to use",bg="#bf360c",font=("Candara",50),fg="white")
selLabel.place(x=250,y=200,width=730)

# Frame visframe1 for t-Distribution
labelFile = tk.Label(visframe1,text="t-Distribution", font=("Helvetica", 20),bg="#d84315",fg="#fff3e0").place(x=0,y=10,width=720)
tk.Label(visframe1,text="t-Value",bg="#d84315",font=("Gadugi",10,"bold"),fg="#fff3e0").place(x=150,y=80,width = 150)
tdistvalue = tk.Entry(visframe1)
tdistvalue.place(x=300, y=80,width = 200)
tk.Label(visframe1,text="Degree of Freedom",bg="#d84315",font=("Gadugi",10,"bold"),fg="#fff3e0").place(x=150,y=120,width = 150)
tdisdfvalue = tk.Entry(visframe1)
tdisdfvalue.place(x=300,y=120,width = 200)
btnCorr = tk.Button(visframe1, text='Calculate', command=calTDis).place(x=325,y=160,width = 150)
tk.Label(visframe1,text="Net Prob(Simpson)",bg="#d84315",font=("Gadugi",10,"bold"),fg="#fff3e0").place(x=150,y=210,width = 150)
txtdissr = StringVar()
tdissr = tk.Entry(visframe1,state=DISABLED,textvariable=txtdissr)
tdissr.place(x=300,y=210,width = 200)
tk.Label(visframe1,text="Net Prob(Mid Point)",bg="#d84315",font=("Gadugi",10,"bold"),fg="#fff3e0").place(x=150,y=260,width = 150)
tnetprob = StringVar()
tdismp = tk.Entry(visframe1,state=DISABLED, textvariable=tnetprob)
tdismp.place(x=300,y=260,width = 200)
tk.Label(visframe1,text="Prob Difference",bg="#d84315",font=("Gadugi",10,"bold"),fg="#fff3e0").place(x=150,y=310,width = 150)
tprobdiff = StringVar()
tdispd = tk.Entry(visframe1,state=DISABLED, textvariable=tprobdiff)
tdispd.place(x=300,y=310,width = 200)


# Frame visframe2 for Chi-square Distribution

labelFile2 = tk.Label(visframe2,text="Chi-Square Distribution", font=("Helvetica", 20),bg=visiframe2bg,fg="#fff3e0").place(x=0,y=10,width=720)
tk.Label(visframe2,text="Chi-Value",bg=visiframe2bg,font=("Gadugi",10,"bold"),fg="#fff3e0").place(x=150,y=80,width = 150)
tdistvalue2 = tk.Entry(visframe2)
tdistvalue2.place(x=300,y=80,width = 200)
tk.Label(visframe2,text="Degree of Freedom",bg=visiframe2bg,font=("Gadugi",10,"bold"),fg="#fff3e0").place(x=150,y=120,width = 150)
tdisdfvalue2 = tk.Entry(visframe2)
tdisdfvalue2.place(x=300,y=120,width = 200)
btnCorr2 = tk.Button(visframe2, text='Calculate', command=calChiDis).place(x=325,y=160,width = 150)
tk.Label(visframe2,text="Net Prob(Simpson)",bg=visiframe2bg,font=("Gadugi",10,"bold"),fg="#fff3e0").place(x=150,y=210,width = 150)
x2 = StringVar()
y2 = StringVar()
z2 = StringVar()
tdissr2 = tk.Entry(visframe2,state=DISABLED, textvariable=x2)
tdissr2.place(x=300,y=210,width = 200)
tk.Label(visframe2,text="Net Prob(Mid Point)",bg=visiframe2bg,font=("Gadugi",10,"bold"),fg="#fff3e0").place(x=150,y=260,width = 150)
tdismp2 = tk.Entry(visframe2,state=DISABLED, textvariable=y2)
tdismp2.place(x=300,y=260,width = 200)
tk.Label(visframe2,text="Prob Difference",bg=visiframe2bg,font=("Gadugi",10,"bold"),fg="#fff3e0").place(x=150,y=310,width = 150)
tdispd2 = tk.Entry(visframe2,state=DISABLED, textvariable=z2)
tdispd2.place(x=300,y=310,width = 200)



# Frame visframe3 for Normal Distribution
labelFile3 = tk.Label(visframe3,text="Normal Distribution", font=("Helvetica", 20),bg=visiframe3bg,fg="#fff3e0").place(x=0,y=10,width=720)
tk.Label(visframe3,text="Enter Z Value",bg=visiframe3bg,font=("Gadugi",10,"bold"),fg="#fff3e0").place(x=150,y=80,width = 150)
tdistvalue3 = tk.Entry(visframe3)
tdistvalue3.place(x=300,y=80,width = 200)
# tk.Label(visframe3,text="Degree of Freedom",bg=visiframe2bg).place(x=150,y=120,width = 150)
# tdisdfvalue3 = tk.Entry(visframe3).place(x=300,y=120,width = 200)
btnCorr3 = tk.Button(visframe3, text='Calculate', command=calNorDis).place(x=325,y=130,width = 150)
tk.Label(visframe3,text="Net Prob(Simpson)",bg=visiframe3bg,font=("Gadugi",10,"bold"),fg="#fff3e0").place(x=150,y=210,width = 150)
x3 = StringVar()
y3 = StringVar()
z3 = StringVar()
tdissr3 = tk.Entry(visframe3,state=DISABLED, textvariable=x3)
tdissr3.place(x=300,y=210,width = 200)
tk.Label(visframe3,text="Net Prob(Mid Point)",bg=visiframe3bg,font=("Gadugi",10,"bold"),fg="#fff3e0").place(x=150,y=260,width = 150)
tdismp3 = tk.Entry(visframe3,state=DISABLED, textvariable=y3)
tdismp3.place(x=300,y=260,width = 200)
tk.Label(visframe3,text="Prob Difference",bg=visiframe3bg,font=("Gadugi",10,"bold"),fg="#fff3e0").place(x=150,y=310,width = 150)
tdispd3 = tk.Entry(visframe3,state=DISABLED, textvariable=z3)
tdispd3.place(x=300,y=310,width = 200)


# Quit Button
quitBt = Button(Frame2, text='Quit', command=root.destroy).place(x=850, y=450, width=100)
TitleFrame.place(x=0,y=0,width=1000,height=500)

mainloop()
